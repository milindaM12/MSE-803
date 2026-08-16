
import sys
import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------
# PATH HANDLING (portable - no hardcoded sandbox paths)
# ---------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

RAW_PATH = sys.argv[1] if len(sys.argv) > 1 else os.path.join(SCRIPT_DIR, "people_data.csv")
OUT_PATH = sys.argv[2] if len(sys.argv) > 2 else os.path.join(SCRIPT_DIR, "data_analysis_results.xlsx")

WORD_NUMBERS = {
    "thirty-eight": 38,
    "thirty eight": 38,
    "sixty five thousand": 65000,
    "sixty-five thousand": 65000,
}

COUNTRY_ALIASES = {"AU": "AUS", "AUS": "AUS", "NZ": "NZ"}


# ---------------------------------------------------------------------
# STAGE 1: CLEANING
# ---------------------------------------------------------------------

def parse_numeric_field(value):
    if pd.isna(value):
        return np.nan
    text = str(value).strip()
    if text == "":
        return np.nan
    lowered = text.lower()
    if lowered in WORD_NUMBERS:
        return float(WORD_NUMBERS[lowered])
    cleaned = text.replace(",", "")
    try:
        return float(cleaned)
    except ValueError:
        return np.nan


def parse_join_date(value):
    if pd.isna(value):
        return pd.NaT
    text = str(value).strip()
    if text == "":
        return pd.NaT
    parsed = pd.to_datetime(text, format="%d/%m/%Y", errors="coerce")
    if pd.isna(parsed):
        parsed = pd.to_datetime(text, format="%Y-%m-%d", errors="coerce")
    if pd.isna(parsed):
        parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    return parsed


def clean_data(raw: pd.DataFrame):
    df = raw.copy()
    quality_report = {}

    quality_report["missing_before"] = (
        df.isna() | df.astype(str).apply(lambda c: c.str.strip() == "")
    ).sum().to_dict()
    quality_report["rows_before"] = len(df)

    df = df.replace(r"^\s*$", np.nan, regex=True)

    def coalesce_group(group: pd.DataFrame) -> pd.Series:
        return group.bfill().ffill().iloc[0]

    duplicate_ids = df["ID"].dropna()
    duplicate_ids = duplicate_ids[duplicate_ids.duplicated()].unique().tolist()
    quality_report["duplicate_ids_merged"] = duplicate_ids

    df_with_id = df[df["ID"].notna()].copy()
    df_no_id = df[df["ID"].isna()].copy()
    df_with_id = df_with_id.groupby("ID", as_index=False, sort=False).apply(
        coalesce_group
    ).reset_index(drop=True)
    df = pd.concat([df_with_id, df_no_id], ignore_index=True)

    df["Name"] = df["Name"].where(df["Name"].notna(), "Unknown")
    df["Age"] = df["Age"].apply(parse_numeric_field)
    df["Net worth"] = df["Net worth"].apply(parse_numeric_field)
    df["Salary"] = df["Salary"].apply(parse_numeric_field)

    df["Country"] = (
        df["Country"].astype(str).str.strip().str.upper().replace("NAN", np.nan)
    )
    df["Country"] = df["Country"].map(COUNTRY_ALIASES).fillna(df["Country"])
    df["Country"] = df["Country"].fillna("Unknown")

    df["Join Date"] = df["Join Date"].apply(parse_join_date)

    quality_report["rows_after"] = len(df)
    quality_report["missing_after"] = df.isna().sum().to_dict()

    return df, quality_report


# ---------------------------------------------------------------------
# STAGE 2: WRITE THE EXCEL WORKBOOK (formula-driven)
# ---------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True)
BODY_FONT = Font(name="Arial")
TITLE_FONT = Font(name="Arial", bold=True, size=13)
MONEY_FMT = "$#,##0.00"
NUM_FMT = "#,##0.00"
DATE_FMT = "DD/MM/YYYY"


def style_header_row(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_cleaned_data_sheet(wb, df: pd.DataFrame):
    ws = wb.active
    ws.title = "Cleaned Data"
    headers = list(df.columns)
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    money_cols = {"Net worth", "Salary"}
    date_cols = {"Join Date"}

    for _, row in df.iterrows():
        values = []
        for col in headers:
            v = row[col]
            if col in date_cols:
                values.append(v.to_pydatetime() if pd.notna(v) else None)
            elif isinstance(v, float) and np.isnan(v):
                values.append(None)
            else:
                values.append(v)
        ws.append(values)

    n_rows = len(df) + 1
    for r in range(2, n_rows + 1):
        for c, col in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            if col in money_cols:
                cell.number_format = MONEY_FMT
            elif col in date_cols:
                cell.number_format = DATE_FMT
            elif col == "Age":
                cell.number_format = "0"

    autosize(ws, [8, 14, 8, 14, 10, 12, 14])
    ws.freeze_panes = "A2"
    return n_rows  # last data row number (including header)


def col_letter_for(headers, name):
    return get_column_letter(headers.index(name) + 1)


def write_descriptive_stats_sheet(wb, headers, last_row):
    ws = wb.create_sheet("Descriptive Stats")
    ws["A1"] = "Descriptive Statistics (live formulas referencing 'Cleaned Data')"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    metrics = [
        "Count (non-missing)",
        "Missing",
        "Mean",
        "Median",
        "Std. Deviation",
        "Min",
        "Max",
        "Range",
        "Q1",
        "Q3",
        "IQR (Q3-Q1)",
        "Lower Outlier Bound",
        "Upper Outlier Bound",
        "Outlier Count",
    ]
    cols = ["Age", "Net worth", "Salary"]

    header_row = 3
    ws.cell(row=header_row, column=1, value="Metric")
    for j, col in enumerate(cols, start=2):
        ws.cell(row=header_row, column=j, value=col)
    style_header_row(ws, header_row, len(cols) + 1)

    money_rows = set()  # rows that hold currency for Net worth/Salary columns

    for i, metric in enumerate(metrics):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=metric).font = BODY_FONT

    def rng(col_name):
        letter = col_letter_for(headers, col_name)
        return f"'Cleaned Data'!{letter}2:{letter}{last_row}"

    for j, col in enumerate(cols, start=2):
        rng_ref = rng(col)
        r = header_row  # will offset below
        ws.cell(row=header_row + 1, column=j, value=f"=COUNT({rng_ref})")
        ws.cell(row=header_row + 2, column=j, value=f"=({last_row}-1)-COUNT({rng_ref})")
        ws.cell(row=header_row + 3, column=j, value=f"=AVERAGE({rng_ref})")
        ws.cell(row=header_row + 4, column=j, value=f"=MEDIAN({rng_ref})")
        ws.cell(row=header_row + 5, column=j, value=f"=STDEV({rng_ref})")
        ws.cell(row=header_row + 6, column=j, value=f"=MIN({rng_ref})")
        ws.cell(row=header_row + 7, column=j, value=f"=MAX({rng_ref})")
        ws.cell(row=header_row + 8, column=j, value=f"=MAX({rng_ref})-MIN({rng_ref})")
        q1_cell = f"{get_column_letter(j)}{header_row + 9}"
        q3_cell = f"{get_column_letter(j)}{header_row + 10}"
        ws.cell(row=header_row + 9, column=j, value=f"=QUARTILE({rng_ref},1)")
        ws.cell(row=header_row + 10, column=j, value=f"=QUARTILE({rng_ref},3)")
        ws.cell(row=header_row + 11, column=j, value=f"={q3_cell}-{q1_cell}")
        iqr_cell = f"{get_column_letter(j)}{header_row + 11}"
        lower_cell = f"{get_column_letter(j)}{header_row + 12}"
        upper_cell = f"{get_column_letter(j)}{header_row + 13}"
        ws.cell(row=header_row + 12, column=j, value=f"={q1_cell}-1.5*{iqr_cell}")
        ws.cell(row=header_row + 13, column=j, value=f"={q3_cell}+1.5*{iqr_cell}")
        ws.cell(
            row=header_row + 14,
            column=j,
            value=f'=COUNTIFS({rng_ref},"<"&{lower_cell})+COUNTIFS({rng_ref},">"&{upper_cell})',
        )

        for i in range(14):
            cell = ws.cell(row=header_row + 1 + i, column=j)
            cell.font = BODY_FONT
            if col in ("Net worth", "Salary") and metrics[i] not in (
                "Count (non-missing)", "Missing", "Outlier Count"
            ):
                cell.number_format = MONEY_FMT
            elif metrics[i] not in ("Count (non-missing)", "Missing", "Outlier Count"):
                cell.number_format = NUM_FMT

    autosize(ws, [22, 16, 16, 16])
    return ws


def write_by_country_sheet(wb, headers, last_row, countries):
    ws = wb.create_sheet("By Country")
    ws["A1"] = "Averages by Country (live formulas referencing 'Cleaned Data')"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    header_row = 3
    cols_out = ["Country", "N (people)", "Avg Age", "Avg Net worth", "Avg Salary"]
    for j, h in enumerate(cols_out, start=1):
        ws.cell(row=header_row, column=j, value=h)
    style_header_row(ws, header_row, len(cols_out))

    country_letter = col_letter_for(headers, "Country")
    country_rng = f"'Cleaned Data'!{country_letter}2:{country_letter}{last_row}"

    for i, country in enumerate(countries):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=country).font = BODY_FONT
        ws.cell(row=r, column=2, value=f'=COUNTIF({country_rng},A{r})').font = BODY_FONT
        for j, col in [(3, "Age"), (4, "Net worth"), (5, "Salary")]:
            col_letter = col_letter_for(headers, col)
            col_rng = f"'Cleaned Data'!{col_letter}2:{col_letter}{last_row}"
            cell = ws.cell(row=r, column=j, value=f'=IFERROR(AVERAGEIF({country_rng},A{r},{col_rng}),"n/a")')
            cell.font = BODY_FONT
            if col in ("Net worth", "Salary"):
                cell.number_format = MONEY_FMT
            else:
                cell.number_format = NUM_FMT

    autosize(ws, [14, 12, 12, 16, 14])
    return ws


def write_correlation_sheet(wb, headers, last_row):
    ws = wb.create_sheet("Correlation")
    ws["A1"] = "Pearson Correlation Matrix (live CORREL formulas)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    cols = ["Age", "Net worth", "Salary"]
    header_row = 3
    ws.cell(row=header_row, column=1, value="")
    for j, col in enumerate(cols, start=2):
        ws.cell(row=header_row, column=j, value=col)
    style_header_row(ws, header_row, len(cols) + 1)

    for i, row_col in enumerate(cols):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=row_col).font = BODY_FONT
        row_letter = col_letter_for(headers, row_col)
        row_rng = f"'Cleaned Data'!{row_letter}2:{row_letter}{last_row}"
        for j, col2 in enumerate(cols, start=2):
            col_letter = col_letter_for(headers, col2)
            col_rng = f"'Cleaned Data'!{col_letter}2:{col_letter}{last_row}"
            cell = ws.cell(row=r, column=j, value=f"=CORREL({row_rng},{col_rng})")
            cell.font = BODY_FONT
            cell.number_format = "0.000"

    autosize(ws, [14, 12, 12, 12])
    return ws


def write_data_quality_sheet(wb, quality_report):
    ws = wb.create_sheet("Data Quality")
    ws["A1"] = "Data Quality Summary (measured during Python cleaning step)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")

    ws["A3"] = "Rows before cleaning"
    ws["B3"] = quality_report["rows_before"]
    ws["A4"] = "Rows after cleaning"
    ws["B4"] = quality_report["rows_after"]
    ws["A5"] = "Duplicate ID(s) merged"
    ws["B5"] = ", ".join(str(x) for x in quality_report["duplicate_ids_merged"]) or "None"

    header_row = 7
    ws.cell(row=header_row, column=1, value="Column")
    ws.cell(row=header_row, column=2, value="Missing BEFORE cleaning")
    ws.cell(row=header_row, column=3, value="Missing AFTER cleaning")
    style_header_row(ws, header_row, 3)

    cols = list(quality_report["missing_before"].keys())
    for i, col in enumerate(cols):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=col).font = BODY_FONT
        ws.cell(row=r, column=2, value=quality_report["missing_before"].get(col, 0)).font = BODY_FONT
        ws.cell(row=r, column=3, value=quality_report["missing_after"].get(col, 0)).font = BODY_FONT

    for row in (3, 4, 5):
        ws.cell(row=row, column=1).font = BODY_FONT
        ws.cell(row=row, column=2).font = BODY_FONT

    autosize(ws, [22, 22, 22])
    return ws


def build_workbook(df, quality_report):
    wb = Workbook()
    headers = list(df.columns)
    last_row = write_cleaned_data_sheet(wb, df)
    write_descriptive_stats_sheet(wb, headers, last_row)
    countries = sorted(df["Country"].dropna().unique().tolist())
    write_by_country_sheet(wb, headers, last_row, countries)
    write_correlation_sheet(wb, headers, last_row)
    write_data_quality_sheet(wb, quality_report)
    return wb


if __name__ == "__main__":
    if not os.path.exists(RAW_PATH):
        print(f"ERROR: could not find input file at:\n  {RAW_PATH}")
        print("Pass the path explicitly, e.g.:")
        print('  python activity1.py "C:\\path\\to\\people_data.csv"')
        sys.exit(1)

    raw_df = pd.read_csv(RAW_PATH, dtype=str)
    clean_df, quality_report = clean_data(raw_df)

    workbook = build_workbook(clean_df, quality_report)
    workbook.save(OUT_PATH)

    print(f"Done. Workbook written to:\n  {OUT_PATH}")
    print("Sheets: Cleaned Data | Descriptive Stats | By Country | Correlation | Data Quality")
