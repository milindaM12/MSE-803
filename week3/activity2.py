"""
activity2.py
============
Continuation of W3-A1 (activity1.py): after cleaning people_data.csv,
this script tries to RECOVER missing numeric values using regression,
comparing a LINEAR model against a NON-LINEAR (polynomial, degree=2)
model in the same style as the lecture's sample code
(nonlinear regression -sample code.py).

USAGE
-----
    python activity2.py
Or:
    python activity2.py "path\\to\\people_data.csv" "path\\to\\data_analysis_results.xlsx"

Looks for people_data.csv next to this script by default, and writes /
overwrites data_analysis_results.xlsx next to it, ADDING to the sheets
produced by activity1.py rather than replacing that work:
    1. Cleaned Data
    2. Descriptive Stats
    3. By Country
    4. Correlation
    5. Data Quality
    6. Missing Value Prediction   <-- NEW: linear vs polynomial comparison
       (with an embedded chart image)

--------------------------------------------------------------------
WHY REGRESSION FOR MISSING VALUES?
--------------------------------------------------------------------
Mean/median imputation (the usual quick fix) replaces every missing
value with the same single number, ignoring everything else we know
about that row. Regression imputation instead asks: "based on this
person's OTHER known attributes, what value would we expect?" - using
the relationship between variables that already exists in the rest of
the dataset.

That only works when the row with the missing value still has at
least one other USABLE predictor. In this dataset:

    - David  is missing Net worth, but his Age (38) and Salary (68000)
              are both known -> a predictor is available -> regression
              imputation is possible.
    - Heidi  is missing Age, Net worth, AND Salary all at once -> there
              is no numeric predictor left for her -> regression
              imputation is NOT possible for her from this dataset
              alone (explained further in the script's output).

--------------------------------------------------------------------
WHY COMPARE LINEAR vs POLYNOMIAL WITH CROSS-VALIDATION?
--------------------------------------------------------------------
With only 7 usable training points (Age & Net worth both known), a
degree-2 polynomial has 3 free parameters vs. 2 for a straight line.
More parameters will ALWAYS fit the training points at least as well,
often better - that is not evidence the curve is "truer", it can
simply mean it is bending itself around noise (overfitting).

The fair test is how well each model predicts a point it did NOT see
during training. With so little data, a single train/test split would
throw away most of it, so this script uses Leave-One-Out Cross-
Validation (LOOCV): fit on 6 points, predict the 7th, repeat for every
point, then compare the resulting out-of-sample errors. Whichever
model generalises better (lower LOOCV error) is used to impute David's
missing Net worth.
"""

import sys
import os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")  # no GUI needed - we save the chart to a file
import matplotlib.pyplot as plt

from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.model_selection import LeaveOneOut

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ---------------------------------------------------------------------
# PATHS (portable)
# ---------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_PATH = sys.argv[1] if len(sys.argv) > 1 else os.path.join(SCRIPT_DIR, "people_data.csv")
OUT_PATH = sys.argv[2] if len(sys.argv) > 2 else os.path.join(SCRIPT_DIR, "data_analysis_results2.xlsx")
CHART_PATH = os.path.join(SCRIPT_DIR, "_regression_comparison_chart.png")

WORD_NUMBERS = {
    "thirty-eight": 38, "thirty eight": 38,
    "sixty five thousand": 65000, "sixty-five thousand": 65000,
}
COUNTRY_ALIASES = {"AU": "AUS", "AUS": "AUS", "NZ": "NZ"}

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True)
BODY_FONT = Font(name="Arial")
TITLE_FONT = Font(name="Arial", bold=True, size=13)
WRAP_FONT_ALIGN = Alignment(wrap_text=True, vertical="top")
MONEY_FMT = "$#,##0.00"
NUM_FMT = "#,##0.00"
DATE_FMT = "DD/MM/YYYY"


# ---------------------------------------------------------------------
# STAGE 1: CLEANING
# ---------------------------------------------------------------------

def parse_numeric_field(value):
    if pd.isna(value):
        return np.nan
    text = str(value).strip()
    if text == "":
        return np.nan
    lowered = text.lower()
    if lowered in WORD_NUMBERS:
        return float(WORD_NUMBERS[lowered])
    cleaned = text.replace(",", "")
    try:
        return float(cleaned)
    except ValueError:
        return np.nan


def parse_join_date(value):
    if pd.isna(value):
        return pd.NaT
    text = str(value).strip()
    if text == "":
        return pd.NaT
    parsed = pd.to_datetime(text, format="%d/%m/%Y", errors="coerce")
    if pd.isna(parsed):
        parsed = pd.to_datetime(text, format="%Y-%m-%d", errors="coerce")
    if pd.isna(parsed):
        parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    return parsed


def clean_data(raw: pd.DataFrame):
    df = raw.copy()
    quality_report = {}
    quality_report["missing_before"] = (
        df.isna() | df.astype(str).apply(lambda c: c.str.strip() == "")
    ).sum().to_dict()
    quality_report["rows_before"] = len(df)

    df = df.replace(r"^\s*$", np.nan, regex=True)

    def coalesce_group(group: pd.DataFrame) -> pd.Series:
        return group.bfill().ffill().iloc[0]

    duplicate_ids = df["ID"].dropna()
    duplicate_ids = duplicate_ids[duplicate_ids.duplicated()].unique().tolist()
    quality_report["duplicate_ids_merged"] = duplicate_ids

    df_with_id = df[df["ID"].notna()].copy()
    df_no_id = df[df["ID"].isna()].copy()
    df_with_id = df_with_id.groupby("ID", as_index=False, sort=False).apply(
        coalesce_group
    ).reset_index(drop=True)
    df = pd.concat([df_with_id, df_no_id], ignore_index=True)

    df["Name"] = df["Name"].where(df["Name"].notna(), "Unknown")
    df["Age"] = df["Age"].apply(parse_numeric_field)
    df["Net worth"] = df["Net worth"].apply(parse_numeric_field)
    df["Salary"] = df["Salary"].apply(parse_numeric_field)

    df["Country"] = df["Country"].astype(str).str.strip().str.upper().replace("NAN", np.nan)
    df["Country"] = df["Country"].map(COUNTRY_ALIASES).fillna(df["Country"])
    df["Country"] = df["Country"].fillna("Unknown")

    df["Join Date"] = df["Join Date"].apply(parse_join_date)

    quality_report["rows_after"] = len(df)
    quality_report["missing_after"] = df.isna().sum().to_dict()
    return df, quality_report


# ---------------------------------------------------------------------
# STAGE 2: REGRESSION-BASED MISSING VALUE PREDICTION
# ---------------------------------------------------------------------

def loocv_evaluate(X, y, degree):
    """
    Leave-One-Out Cross-Validation for either a linear (degree=1) or
    polynomial (degree=2) model. Returns (cv_mse, cv_r2, out-of-sample
    predictions aligned to y's original order).
    """
    loo = LeaveOneOut()
    preds = np.zeros_like(y, dtype=float)

    for train_idx, test_idx in loo.split(X):
        X_train, X_test = X[train_idx], X[test_idx]
        y_train = y[train_idx]

        if degree == 1:
            model = LinearRegression()
            model.fit(X_train, y_train)
            pred = model.predict(X_test)
        else:
            poly = PolynomialFeatures(degree=degree, include_bias=False)
            X_train_p = poly.fit_transform(X_train)
            X_test_p = poly.transform(X_test)
            model = LinearRegression()
            model.fit(X_train_p, y_train)
            pred = model.predict(X_test_p)

        preds[test_idx] = pred.ravel()

    cv_mse = mean_squared_error(y, preds)
    cv_r2 = r2_score(y, preds)
    return cv_mse, cv_r2, preds


def fit_full_models(X, y):
    """Fit both models on ALL available training rows (in-sample)."""
    linear_model = LinearRegression().fit(X, y)
    y_pred_lin = linear_model.predict(X)
    in_mse_lin = mean_squared_error(y, y_pred_lin)
    in_r2_lin = r2_score(y, y_pred_lin)

    poly = PolynomialFeatures(degree=2, include_bias=False)
    X_poly = poly.fit_transform(X)
    poly_model = LinearRegression().fit(X_poly, y)
    y_pred_poly = poly_model.predict(X_poly)
    in_mse_poly = mean_squared_error(y, y_pred_poly)
    in_r2_poly = r2_score(y, y_pred_poly)

    return {
        "linear_model": linear_model,
        "poly": poly,
        "poly_model": poly_model,
        "in_mse_lin": in_mse_lin, "in_r2_lin": in_r2_lin,
        "in_mse_poly": in_mse_poly, "in_r2_poly": in_r2_poly,
    }


def run_imputation(df: pd.DataFrame):
    """
    Predict David's missing Net worth from Age, using Linear vs
    Polynomial(deg=2) regression, chosen via LOOCV. Also documents why
    Heidi's missing values cannot be regressed from this dataset.
    """
    target_col, predictor_col = "Net worth", "Age"

    train_mask = df[predictor_col].notna() & df[target_col].notna()
    train_df = df[train_mask]
    X = train_df[[predictor_col]].to_numpy(dtype=float)
    y = train_df[target_col].to_numpy(dtype=float)

    missing_mask = df[target_col].isna() & df[predictor_col].notna()
    predict_rows = df[missing_mask]

    full = fit_full_models(X, y)
    cv_mse_lin, cv_r2_lin, _ = loocv_evaluate(X, y, degree=1)
    cv_mse_poly, cv_r2_poly, _ = loocv_evaluate(X, y, degree=2)

    better_model = "Polynomial (degree 2)" if cv_mse_poly < cv_mse_lin else "Linear"

    predictions = []
    for _, row in predict_rows.iterrows():
        x_val = np.array([[row[predictor_col]]], dtype=float)
        lin_pred = float(full["linear_model"].predict(x_val)[0])
        poly_pred = float(full["poly_model"].predict(full["poly"].transform(x_val))[0])
        chosen = poly_pred if better_model.startswith("Polynomial") else lin_pred
        predictions.append({
            "Name": row["Name"], "ID": row["ID"], predictor_col: row[predictor_col],
            "linear_pred": lin_pred, "poly_pred": poly_pred, "chosen": chosen,
        })

    unresolved = df[df[["Age", "Net worth", "Salary"]].isna().all(axis=1)]

    results = {
        "target_col": target_col, "predictor_col": predictor_col,
        "n_train": len(train_df), "X": X, "y": y,
        "in_mse_lin": full["in_mse_lin"], "in_r2_lin": full["in_r2_lin"],
        "in_mse_poly": full["in_mse_poly"], "in_r2_poly": full["in_r2_poly"],
        "cv_mse_lin": cv_mse_lin, "cv_r2_lin": cv_r2_lin,
        "cv_mse_poly": cv_mse_poly, "cv_r2_poly": cv_r2_poly,
        "better_model": better_model,
        "predictions": predictions,
        "linear_model": full["linear_model"], "poly": full["poly"], "poly_model": full["poly_model"],
        "unresolved_names": unresolved["Name"].tolist(),
    }
    return results


def make_comparison_chart(results, path):
    X, y = results["X"], results["y"]
    x_min, x_max = X.min() - 3, X.max() + 3
    x_line = np.linspace(x_min, x_max, 200).reshape(-1, 1)

    y_lin = results["linear_model"].predict(x_line)
    y_poly = results["poly_model"].predict(results["poly"].transform(x_line))

    fig, ax = plt.subplots(figsize=(8, 5.5), dpi=140)
    ax.scatter(X, y, color="#2b5fad", s=55, zorder=3, label="Known data (Age vs Net worth)")
    ax.plot(x_line, y_lin, color="#d94f4f", linewidth=2, label="Linear fit")
    ax.plot(x_line, y_poly, color="#2fa66a", linewidth=2, linestyle="--", label="Polynomial fit (deg=2)")

    for p in results["predictions"]:
        ax.scatter(p["Age"], p["linear_pred"], marker="^", s=140, color="#d94f4f",
                    edgecolor="black", zorder=4,
                    label=f"Linear prediction ({p['Name']})")
        ax.scatter(p["Age"], p["poly_pred"], marker="D", s=110, color="#2fa66a",
                    edgecolor="black", zorder=4,
                    label=f"Polynomial prediction ({p['Name']})")

    ax.set_xlabel("Age")
    ax.set_ylabel("Net worth")
    ax.set_title("Predicting Missing Net Worth from Age:\nLinear vs Polynomial (degree 2) Regression")
    ax.legend(fontsize=8, loc="upper left")
    ax.grid(alpha=0.25)
    fig.tight_layout()
    fig.savefig(path)
    plt.close(fig)


# ---------------------------------------------------------------------
# EXCEL WRITING
# ---------------------------------------------------------------------

def style_header_row(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_cleaned_data_sheet(wb, df):
    ws = wb.active
    ws.title = "Cleaned Data"
    headers = list(df.columns)
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    money_cols, date_cols = {"Net worth", "Salary"}, {"Join Date"}
    for _, row in df.iterrows():
        values = []
        for col in headers:
            v = row[col]
            if col in date_cols:
                values.append(v.to_pydatetime() if pd.notna(v) else None)
            elif isinstance(v, float) and np.isnan(v):
                values.append(None)
            else:
                values.append(v)
        ws.append(values)
    n_rows = len(df) + 1
    for r in range(2, n_rows + 1):
        for c, col in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            if col in money_cols:
                cell.number_format = MONEY_FMT
            elif col in date_cols:
                cell.number_format = DATE_FMT
            elif col == "Age":
                cell.number_format = "0"
    autosize(ws, [8, 14, 8, 14, 10, 12, 14])
    ws.freeze_panes = "A2"
    return n_rows


def col_letter_for(headers, name):
    return get_column_letter(headers.index(name) + 1)


def write_descriptive_stats_sheet(wb, headers, last_row):
    ws = wb.create_sheet("Descriptive Stats")
    ws["A1"] = "Descriptive Statistics (live formulas referencing 'Cleaned Data')"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    metrics = ["Count (non-missing)", "Missing", "Mean", "Median", "Std. Deviation",
               "Min", "Max", "Range", "Q1", "Q3", "IQR (Q3-Q1)",
               "Lower Outlier Bound", "Upper Outlier Bound", "Outlier Count"]
    cols = ["Age", "Net worth", "Salary"]
    header_row = 3
    ws.cell(row=header_row, column=1, value="Metric")
    for j, col in enumerate(cols, start=2):
        ws.cell(row=header_row, column=j, value=col)
    style_header_row(ws, header_row, len(cols) + 1)
    for i, metric in enumerate(metrics):
        ws.cell(row=header_row + 1 + i, column=1, value=metric).font = BODY_FONT

    def rng(col_name):
        letter = col_letter_for(headers, col_name)
        return f"'Cleaned Data'!{letter}2:{letter}{last_row}"

    for j, col in enumerate(cols, start=2):
        rng_ref = rng(col)
        ws.cell(row=header_row + 1, column=j, value=f"=COUNT({rng_ref})")
        ws.cell(row=header_row + 2, column=j, value=f"=({last_row}-1)-COUNT({rng_ref})")
        ws.cell(row=header_row + 3, column=j, value=f"=AVERAGE({rng_ref})")
        ws.cell(row=header_row + 4, column=j, value=f"=MEDIAN({rng_ref})")
        ws.cell(row=header_row + 5, column=j, value=f"=STDEV({rng_ref})")
        ws.cell(row=header_row + 6, column=j, value=f"=MIN({rng_ref})")
        ws.cell(row=header_row + 7, column=j, value=f"=MAX({rng_ref})")
        ws.cell(row=header_row + 8, column=j, value=f"=MAX({rng_ref})-MIN({rng_ref})")
        q1_cell = f"{get_column_letter(j)}{header_row + 9}"
        q3_cell = f"{get_column_letter(j)}{header_row + 10}"
        ws.cell(row=header_row + 9, column=j, value=f"=QUARTILE({rng_ref},1)")
        ws.cell(row=header_row + 10, column=j, value=f"=QUARTILE({rng_ref},3)")
        ws.cell(row=header_row + 11, column=j, value=f"={q3_cell}-{q1_cell}")
        iqr_cell = f"{get_column_letter(j)}{header_row + 11}"
        lower_cell = f"{get_column_letter(j)}{header_row + 12}"
        upper_cell = f"{get_column_letter(j)}{header_row + 13}"
        ws.cell(row=header_row + 12, column=j, value=f"={q1_cell}-1.5*{iqr_cell}")
        ws.cell(row=header_row + 13, column=j, value=f"={q3_cell}+1.5*{iqr_cell}")
        ws.cell(row=header_row + 14, column=j,
                value=f'=COUNTIFS({rng_ref},"<"&{lower_cell})+COUNTIFS({rng_ref},">"&{upper_cell})')
        for i in range(14):
            cell = ws.cell(row=header_row + 1 + i, column=j)
            cell.font = BODY_FONT
            if col in ("Net worth", "Salary") and metrics[i] not in ("Count (non-missing)", "Missing", "Outlier Count"):
                cell.number_format = MONEY_FMT
            elif metrics[i] not in ("Count (non-missing)", "Missing", "Outlier Count"):
                cell.number_format = NUM_FMT
    autosize(ws, [22, 16, 16, 16])


def write_by_country_sheet(wb, headers, last_row, countries):
    ws = wb.create_sheet("By Country")
    ws["A1"] = "Averages by Country (live formulas referencing 'Cleaned Data')"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    header_row = 3
    cols_out = ["Country", "N (people)", "Avg Age", "Avg Net worth", "Avg Salary"]
    for j, h in enumerate(cols_out, start=1):
        ws.cell(row=header_row, column=j, value=h)
    style_header_row(ws, header_row, len(cols_out))
    country_letter = col_letter_for(headers, "Country")
    country_rng = f"'Cleaned Data'!{country_letter}2:{country_letter}{last_row}"
    for i, country in enumerate(countries):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=country).font = BODY_FONT
        ws.cell(row=r, column=2, value=f'=COUNTIF({country_rng},A{r})').font = BODY_FONT
        for j, col in [(3, "Age"), (4, "Net worth"), (5, "Salary")]:
            col_letter = col_letter_for(headers, col)
            col_rng = f"'Cleaned Data'!{col_letter}2:{col_letter}{last_row}"
            cell = ws.cell(row=r, column=j, value=f'=IFERROR(AVERAGEIF({country_rng},A{r},{col_rng}),"n/a")')
            cell.font = BODY_FONT
            cell.number_format = MONEY_FMT if col in ("Net worth", "Salary") else NUM_FMT
    autosize(ws, [14, 12, 12, 16, 14])


def write_correlation_sheet(wb, headers, last_row):
    ws = wb.create_sheet("Correlation")
    ws["A1"] = "Pearson Correlation Matrix (live CORREL formulas)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    cols = ["Age", "Net worth", "Salary"]
    header_row = 3
    ws.cell(row=header_row, column=1, value="")
    for j, col in enumerate(cols, start=2):
        ws.cell(row=header_row, column=j, value=col)
    style_header_row(ws, header_row, len(cols) + 1)
    for i, row_col in enumerate(cols):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=row_col).font = BODY_FONT
        row_letter = col_letter_for(headers, row_col)
        row_rng = f"'Cleaned Data'!{row_letter}2:{row_letter}{last_row}"
        for j, col2 in enumerate(cols, start=2):
            col_letter = col_letter_for(headers, col2)
            col_rng = f"'Cleaned Data'!{col_letter}2:{col_letter}{last_row}"
            cell = ws.cell(row=r, column=j, value=f"=CORREL({row_rng},{col_rng})")
            cell.font = BODY_FONT
            cell.number_format = "0.000"
    autosize(ws, [14, 12, 12, 12])


def write_data_quality_sheet(wb, quality_report):
    ws = wb.create_sheet("Data Quality")
    ws["A1"] = "Data Quality Summary (measured during Python cleaning step)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")
    ws["A3"] = "Rows before cleaning"; ws["B3"] = quality_report["rows_before"]
    ws["A4"] = "Rows after cleaning"; ws["B4"] = quality_report["rows_after"]
    ws["A5"] = "Duplicate ID(s) merged"
    ws["B5"] = ", ".join(str(x) for x in quality_report["duplicate_ids_merged"]) or "None"
    header_row = 7
    ws.cell(row=header_row, column=1, value="Column")
    ws.cell(row=header_row, column=2, value="Missing BEFORE cleaning")
    ws.cell(row=header_row, column=3, value="Missing AFTER cleaning")
    style_header_row(ws, header_row, 3)
    cols = list(quality_report["missing_before"].keys())
    for i, col in enumerate(cols):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=col).font = BODY_FONT
        ws.cell(row=r, column=2, value=quality_report["missing_before"].get(col, 0)).font = BODY_FONT
        ws.cell(row=r, column=3, value=quality_report["missing_after"].get(col, 0)).font = BODY_FONT
    for row in (3, 4, 5):
        ws.cell(row=row, column=1).font = BODY_FONT
        ws.cell(row=row, column=2).font = BODY_FONT
    autosize(ws, [22, 22, 22])


def write_prediction_sheet(wb, results, chart_path):
    ws = wb.create_sheet("Missing Value Prediction")
    ws["A1"] = "Missing Value Prediction: Linear vs Polynomial Regression"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ws["A3"] = f"Target column: {results['target_col']}   |   Predictor: {results['predictor_col']}   |   Training rows used: {results['n_train']}"
    ws["A3"].font = BODY_FONT

    # --- Model comparison table ---
    header_row = 5
    cols = ["Metric", "Linear Regression", "Polynomial Regression (deg=2)"]
    for j, h in enumerate(cols, start=1):
        ws.cell(row=header_row, column=j, value=h)
    style_header_row(ws, header_row, len(cols))

    rows = [
        ("In-sample MSE (fit on all training rows, tested on same rows)",
         results["in_mse_lin"], results["in_mse_poly"]),
        ("In-sample R\u00b2", results["in_r2_lin"], results["in_r2_poly"]),
        ("LOOCV MSE (out-of-sample, fair comparison)",
         results["cv_mse_lin"], results["cv_mse_poly"]),
        ("LOOCV R\u00b2 (out-of-sample)", results["cv_r2_lin"], results["cv_r2_poly"]),
    ]
    for i, (label, v1, v2) in enumerate(rows):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = BODY_FONT
        c1 = ws.cell(row=r, column=2, value=round(v1, 4)); c1.font = BODY_FONT
        c2 = ws.cell(row=r, column=3, value=round(v2, 4)); c2.font = BODY_FONT

    verdict_row = header_row + len(rows) + 2
    ws.cell(row=verdict_row, column=1,
            value=f"Better generalising model (lower LOOCV MSE): {results['better_model']}").font = Font(name="Arial", bold=True)

    # --- Predicted values table ---
    pred_header_row = verdict_row + 2
    pred_cols = ["Name", "ID", "Age (known)", "Linear prediction", "Polynomial prediction", "Chosen value used"]
    for j, h in enumerate(pred_cols, start=1):
        ws.cell(row=pred_header_row, column=j, value=h)
    style_header_row(ws, pred_header_row, len(pred_cols))

    r = pred_header_row
    for i, p in enumerate(results["predictions"]):
        r = pred_header_row + 1 + i
        ws.cell(row=r, column=1, value=p["Name"]).font = BODY_FONT
        ws.cell(row=r, column=2, value=p["ID"]).font = BODY_FONT
        ws.cell(row=r, column=3, value=p["Age"]).font = BODY_FONT
        c = ws.cell(row=r, column=4, value=round(p["linear_pred"], 2)); c.font = BODY_FONT; c.number_format = MONEY_FMT
        c = ws.cell(row=r, column=5, value=round(p["poly_pred"], 2)); c.font = BODY_FONT; c.number_format = MONEY_FMT
        c = ws.cell(row=r, column=6, value=round(p["chosen"], 2)); c.font = Font(name="Arial", bold=True); c.number_format = MONEY_FMT

    # --- Unresolved cases note ---
    note_row = r + 2
    unresolved = ", ".join(results["unresolved_names"]) if results["unresolved_names"] else "None"
    ws.cell(row=note_row, column=1,
            value=f"Could NOT be predicted by regression (no known predictor available): {unresolved}").font = Font(name="Arial", italic=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
    explain_row = note_row + 1
    ws.cell(row=explain_row, column=1,
            value=("Reason: Heidi is missing Age, Net worth, AND Salary simultaneously, so no "
                   "numeric predictor remains for her row. Regression imputation needs at least "
                   "one known variable to predict from; a mean/median fallback (or dropping the "
                   "row from numeric analyses) is the practical alternative for her."))
    ws.cell(row=explain_row, column=1).alignment = WRAP_FONT_ALIGN
    ws.cell(row=explain_row, column=1).font = BODY_FONT
    ws.merge_cells(start_row=explain_row, start_column=1, end_row=explain_row, end_column=6)
    ws.row_dimensions[explain_row].height = 45

    autosize(ws, [12, 8, 12, 18, 22, 18])

    # --- Embed the chart ---
    if os.path.exists(chart_path):
        img = XLImage(chart_path)
        img.width, img.height = 560, 385
        anchor_row = explain_row + 3
        ws.add_image(img, f"A{anchor_row}")

    return ws


def build_workbook(df, quality_report, results, chart_path):
    wb = Workbook()
    headers = list(df.columns)
    last_row = write_cleaned_data_sheet(wb, df)
    write_descriptive_stats_sheet(wb, headers, last_row)
    countries = sorted(df["Country"].dropna().unique().tolist())
    write_by_country_sheet(wb, headers, last_row, countries)
    write_correlation_sheet(wb, headers, last_row)
    write_data_quality_sheet(wb, quality_report)
    write_prediction_sheet(wb, results, chart_path)
    return wb


if __name__ == "__main__":
    if not os.path.exists(RAW_PATH):
        print(f"ERROR: could not find input file at:\n  {RAW_PATH}")
        sys.exit(1)

    raw_df = pd.read_csv(RAW_PATH, dtype=str)
    clean_df, quality_report = clean_data(raw_df)

    results = run_imputation(clean_df)
    make_comparison_chart(results, CHART_PATH)

    print("=" * 70)
    print("MISSING VALUE PREDICTION: LINEAR vs POLYNOMIAL REGRESSION")
    print("=" * 70)
    print(f"Target: {results['target_col']}  |  Predictor: {results['predictor_col']}  "
          f"|  Training rows: {results['n_train']}")
    print(f"\n{'Metric':45s}{'Linear':>15s}{'Polynomial':>15s}")
    print(f"{'In-sample MSE':45s}{results['in_mse_lin']:>15.2f}{results['in_mse_poly']:>15.2f}")
    print(f"{'In-sample R2':45s}{results['in_r2_lin']:>15.4f}{results['in_r2_poly']:>15.4f}")
    print(f"{'LOOCV MSE (out-of-sample)':45s}{results['cv_mse_lin']:>15.2f}{results['cv_mse_poly']:>15.2f}")
    print(f"{'LOOCV R2 (out-of-sample)':45s}{results['cv_r2_lin']:>15.4f}{results['cv_r2_poly']:>15.4f}")
    print(f"\nBetter generalising model: {results['better_model']}")
    for p in results["predictions"]:
        print(f"\n{p['Name']} (Age={p['Age']:.0f}): "
              f"linear pred=${p['linear_pred']:,.2f}, poly pred=${p['poly_pred']:,.2f}, "
              f"chosen=${p['chosen']:,.2f}")
    if results["unresolved_names"]:
        print(f"\nCould not predict (no known predictor available): {', '.join(results['unresolved_names'])}")

    workbook = build_workbook(clean_df, quality_report, results, CHART_PATH)
    workbook.save(OUT_PATH)
    print(f"\nWorkbook written to:\n  {OUT_PATH}")
    print("Sheets: Cleaned Data | Descriptive Stats | By Country | Correlation | "
          "Data Quality | Missing Value Prediction")
